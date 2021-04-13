from pynput import mouse #import Button, Controller
from pynput import keyboard
import time
import openpyxl
import pyperclip

PODROBY_ID = [81291, 81391, 81491, 81591, 84291, 84391, 84491, 87291, 87391, 87491]

def click_kb(controller, key, times=1):
    for i in range(times):
        controller.press(key)
        controller.release(key)

def move_click(controler, pos, interval=0.1):
    controler.position = pos
    time.sleep(interval)
    controler.click(mouse.Button.left)

class Clicker():
    def __init__(self, template_fr_filename, template_other_filename, template_info_filename, dates_fr_filename, dates_ot_filename):
        self.mouse_controller = mouse.Controller()
        self.keyboard_controller = keyboard.Controller()
        self.cursor_x = 235
        self.cursor_y = 355
        self.interval = 0.2
        with open(template_fr_filename, encoding="utf8") as f:
            self.template_lines_fr = f.read().splitlines()
        with open(template_other_filename, encoding="utf8") as f:
            self.template_lines_other = f.read().splitlines()
        with open(template_info_filename, encoding="utf8") as f:
            self.template_lines_info = f.read().splitlines()
        with open(dates_fr_filename, encoding="utf8") as f:
            self.dates_fr = f.read().splitlines()
        with open(dates_ot_filename, encoding="utf8") as f:
            self.dates_ot = f.read().splitlines()

    def read_sheet(self, wb_filename):
        wb = openpyxl.load_workbook(wb_filename)
        self.sheet = wb['swieze']
    
    def insert_fresh(self, id):
        number = 3 if int(id) in PODROBY_ID else 4
        temperature_txt = f"Entreposer entre 0°C - +{number}°C"
        for i, line in enumerate(self.template_lines_other):
            time.sleep(self.interval)
            if i == 1:
                # self.keyboard_controller.type('\b')
                pyperclip.copy(temperature_txt)
                time.sleep(self.interval)
                with self.keyboard_controller.pressed(keyboard.Key.ctrl):
                    self.keyboard_controller.press('v')
                    self.keyboard_controller.release('v')
                time.sleep(self.interval)
            elif i == 2:
                self.keyboard_controller.type('\b')
            else:
                self.keyboard_controller.type(line)
            self.keyboard_controller.type('\t')

    def do_action(self, start, stop):
        for i in range(start, stop): #self.sheet.max_row
            id = self.sheet['A' + str(i)].value
            if((int(id) - 90) % 100 < 4 and (int(id) - 90) % 100 > 0):
                move_click(self.mouse_controller, (self.cursor_x, self.cursor_y))
                time.sleep(self.interval)
                self.keyboard_controller.type(str(id))
                click_kb(self.keyboard_controller, keyboard.Key.enter)
                time.sleep(self.interval)
                
                click_kb(self.keyboard_controller, keyboard.Key.f10, 2)
                
                self.insert_fresh(int(id))
                
                time.sleep(self.interval)

                click_kb(self.keyboard_controller, keyboard.Key.f2)
                print("Zrobiono rzad: " + str(i) + " o indeksie: " + str(id))
                time.sleep(0.5)
        
cl = Clicker('frozen.txt', "rest.txt", "other_info.txt", "dates_fr.txt", "dates_ot.txt")
cl.read_sheet("asort.xlsx")

time.sleep(2)

cl.do_action(483, cl.sheet.max_row) #cl.sheet.max_row